from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import numpy as np

from dataset import *
import argparse
from collections import Counter
import random
import os
import re

from matplotlib.ticker import MaxNLocator, MultipleLocator, FuncFormatter


def all_equal(lst):
    return all(x == lst[0] for x in lst)


def find_most_common_elements(input_list):
    counter = Counter(input_list)
    max_count = max(counter.values())
    most_common_elements = [element for element, count in counter.items() if count == max_count]
    return most_common_elements, max_count


def get_most_common_answer(outputs):
    outputs = [output for output in outputs if output != None]
    if outputs == []:
        return None
    most_common_elements, max_count = find_most_common_elements(outputs)
    most_common_answer = random.choice(most_common_elements)
    return most_common_answer


def get_cost(model_name, prompt_tokens, completion_tokens):
    prompt_tokens = float(prompt_tokens)
    completion_tokens = float(completion_tokens)
    if "gemini" in model_name:
        cost = prompt_tokens * 0.075 + completion_tokens * 0.3
    elif model_name == "gpt-3.5-turbo-0613":
        cost = prompt_tokens * 1.5 + completion_tokens * 2
    elif model_name == "gpt-4o-mini":
        cost = prompt_tokens * 0.15 + completion_tokens * 0.6
    else:
        cost = prompt_tokens * 0.15 + completion_tokens * 0.6
    cost = cost / 10 ** 6
    return cost


if __name__ == "__main__":
    model_names = ["Qwen/Qwen2.5-7B-Instruct",
                   "meta-llama/Meta-Llama-3-8B-Instruct",
                   "THUDM/glm-4-9b-chat",
                   "gemini-1.5-flash",
                   "gpt-4o-mini",
                   "microsoft/Phi-3.5-mini-instruct"]
    
    datasets = ["GSM8K",
                "GPQA",
                "GSM-Hard",
                "MATH",
                "MMLU-high_school_physics",
                "MMLU-high_school_chemistry",
                "MMLU-high_school_biology",
                "AIME_2024"]
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--model_name", type=str, default=model_names[0])
    parser.add_argument("--dataset", type=str, default=datasets[0])
    parser.add_argument("--split", type=str, default="test")
    args = parser.parse_args()
    
    sampling_times = [1, 3, 5, 7, 10, 15]
    pos = - sampling_times[-1] / 30
    
    model_names_formal = {
        "Qwen2.5-7B-Instruct": "Qwen2.5-7B-Instruct",
        "Llama-3-8B-Instruct": "LLaMA-3-8B-Instruct",
        "glm-4-9b-chat": "GLM-4-9B-Chat",
        "gemini-1.5-flash": "Gemini-1.5-Flash",
        "gpt-4o-mini": "GPT-4o-mini",
        "Phi-3.5-mini-instruct": "Phi-3.5-mini-Instruct"
    }

    # Maximum sampling time for each LLM, in order to avoid loading files that are recording the current runnning results.
    N_dict = {
        "Qwen2.5-7B-Instruct": 16,
        "Llama-3-8B-Instruct": 16,
        "glm-4-9b-chat": 16,
        "gemini-1.5-flash": 16,
        "gpt-4o-mini": 16,
        "Phi-3.5-mini-instruct": 16
    }
    
    marker_dict = {
    'DiP': '^',  
    'CoT': '^',  
    'L2M': '^',  
    'ToT': 'o',  
    'S-RF':'o', 
    'SBP': '^',  
    'AnP': '^',  
    'MAD': 'o'   
    }
    

    if "/" in args.model_name:
        args.model_name = args.model_name.split('/')[-1]
    N_max = N_dict[args.model_name]
    
    if not os.path.exists(os.path.join(log_path, args.dataset, args.model_name, "pics")):
        os.mkdir(os.path.join(log_path, args.dataset, args.model_name, "pics"))

    if "/" in args.model_name:
        args.model_name = args.model_name.split('/')[-1]
        
    path = os.path.join(log_path, args.dataset, f"{args.dataset}_N.xlsx")
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        
    unique_labels = []
    
    headers = ["Method", "Subject", "x-shot", "num", "prompt_tokens", "completion_tokens", "cost",  "tokens", "accuracy"]
    
    if args.model_name in wb.sheetnames:
        ws_new = wb[args.model_name]
        ws_new.delete_rows(1, ws_new.max_row)  
        ws_new.delete_cols(1, ws_new.max_column)  
    else:
        ws_new = wb.create_sheet(args.model_name)
        ws_new.append(headers)

    reasonings = ["DiP", "CoT", "L2M", "SBP", "AnP", "S-RF", "ToT_3", "ToT_5", "ToT_10", "MAD"]

    try:    
        model_name = args.model_name
        sheet = wb.get_sheet_by_name(model_name)

        tokens = []
        accuracy = []
        cost = []
        labels = []
        flag = 0
        for row in sheet.iter_rows(values_only=True):
            if flag == 0:
                flag = 1
                continue
            # if row[0] != "tot-io":
            labels.append(row[0])
            tokens.append(int(row[-2]))
            accuracy.append(float(row[-1]))
            cost.append(float(row[-3]))
        counter = Counter(labels)
        assert counter["DiP"] >= N_max
        assert counter["CoT"] >= N_max
        assert counter["L2M"] >= N_max
        assert counter["SBP"] >= N_max
        assert counter["AnP"] >= N_max
        # assert counter["ToT"] >= 3
        # assert counter["S-RF"] > 3
        # assert counter["MAD"] > 3
    except:
        for reasoning in reasonings:
            try:
                if reasoning == "DiP":
                    args.reasoning = "DiP"
                    args.shot = 0
                    for N in range(0, N_max+1):
                        if not os.path.exists(os.path.join(log_path, args.dataset, args.model_name, f"{args.reasoning}_{args.shot}_{N}.json")):
                            break
                    # N = N - 1
                elif reasoning == "CoT":
                    args.reasoning = "CoT"
                    args.shot = 0
                    for N in range(0, N_max+1):
                        if not os.path.exists(os.path.join(log_path, args.dataset, args.model_name, f"{args.reasoning}_{args.shot}_{N}.json")):
                            break
                    # N = N - 1
                elif reasoning == "L2M":
                    args.reasoning = "L2M"
                    if args.dataset in ["GSM8K", "GSM-Hard"]:
                        args.shot = 1
                    else:
                        args.shot = 0
                    for N in range(0, N_max+1):
                        if not os.path.exists(os.path.join(log_path, args.dataset, args.model_name, f"{args.reasoning}_{args.shot}_{N}.json")):
                            break
                    # N = N - 1
                elif reasoning == "ToT_3":
                    args.reasoning = "ToT"
                    args.shot = 3
                    N = 1
                elif reasoning == "ToT_5":
                    args.reasoning = "ToT"
                    args.shot = 5
                    N = 1
                elif reasoning == "ToT_10":
                    args.reasoning = "ToT"
                    args.shot = 10
                    N = 1
                elif reasoning == "S-RF":
                    args.reasoning = reasoning
                    args.shot = 0
                    N = (len(logs[-1]["record"]) - 1) // 2
                    
                elif reasoning == "SBP":
                    args.reasoning = "SBP"
                    args.shot = 0
                    for N in range(0, N_max+1):
                        if not os.path.exists(os.path.join(log_path, args.dataset, args.model_name, f"{args.reasoning}_{args.shot}_{N}.json")):
                            break
                    # N = N - 1
                elif reasoning == "AnP":
                    args.reasoning = "AnP"
                    args.shot = 1
                    for N in range(0, N_max+1):
                        if not os.path.exists(os.path.join(log_path, args.dataset, args.model_name, f"{args.reasoning}_{args.shot}_{N}.json")):
                            break
                elif reasoning == "MAD":
                    args.reasoning = "MAD"
                    args.shot = 0
                    args.n = 0
                    logs = read_logs(args)
                    N = len(logs[0]["record"])
            
                logs_list = []
                assert N != 0
                for m in range(0, N):
                    if args.reasoning in ["DiP", "CoT", "L2M", "AnP", "SBP"]:
                        args.nums = range(0,m+1)
                        args.n = m
                        logs = read_logs(args)
                        logs_list.append(logs)
                        if (m+1) not in sampling_times:
                            continue
                    elif args.reasoning in ["S-RF", "MAD"]:
                        args.n = 0
                        args.nums = range(0, m+1)
                        logs = read_logs(args)
                        logs_list.append(logs)
                    elif args.reasoning == "ToT":
                        args.nums = range(0, m+1)
                        for n in range(0, 5):
                            try:
                                args.n = n
                                logs = read_logs(args)
                                logs_list.append(logs)
                            except:
                                break
                    
                    accs = []
                    prompt_tokens_ = 0
                    completion_tokens_ = 0
                    l = len(logs_list[0])
                    for count in range(0, 5):
                        acc_num = 0
                        if args.dataset in ["GSM8K", "GSM-Hard", "MATH", "AIME_2024"]:
                            subject = "mathematic"
                        for j in range(0, l):
                            if args.dataset in ["GPQA"] or "MMLU" in args.dataset:
                                subject = logs_list[0][j]["subject"]
                            key = logs_list[0][j]['key']
                            if args.reasoning in ["DiP", "CoT", "L2M", "AnP"]:
                                output_keys = [parse_answer(args, log[j]['record']['output']) for log in logs_list]
                                output_keys = [output for output in output_keys if output != None]
                                if len(output_keys) == 0:
                                    output_keys = [None]
                                
                                if count == 0:
                                    prompt_tokens = [log[j]['record']['usage']['prompt_tokens'] for log in  logs_list[:1]]
                                    completion_tokens = [log[j]['record']['usage']['completion_tokens'] for log in  logs_list]

                            elif args.reasoning == "S-RF":
                                output_keys = [parse_answer(args, logs[j]["record"][f"output{m+1}"]["output"])]
                                if count == 0:
                                    prompt_tokens = [logs[j]["record"]["output0"]['usage']["prompt_tokens"]]
                                    completion_tokens = [logs[j]["record"]["output0"]['usage']["completion_tokens"]]
                                    for k in range(0, m+1):
                                        prompt_tokens += [logs[j]["record"][f"problems{k+1}"]['usage']["prompt_tokens"], logs[j]["record"][f"output{k+1}"]['usage']["prompt_tokens"]]
                                        completion_tokens += [logs[j]["record"][f"problems{k+1}"]['usage']["completion_tokens"], logs[j]["record"][f"output{k+1}"]['usage']["completion_tokens"]] 
                                        
                            elif args.reasoning == "ToT":
                                indexes = []
                                for logs in logs_list:
                                    index = parse_best_solution(logs[j]["record"]["choose"]["output"])
                                    if index != None and "0" < index and index <= str(args.shot):
                                        index = int(index) - 1
                                    else:
                                        index = random.choice(range(0, args.shot))
                                    indexes.append(index)
                                solutions = logs_list[0][j]["record"]["solutions"]
                                index = get_most_common_answer(indexes)
                                best_solution = solutions[index]
                                output_keys = [parse_answer(args, best_solution["output"])]
                                if count == 0:
                                    prompt_tokens = []
                                    completion_tokens = []
                                    for k in range(0, len(solutions)):
                                        solution = solutions[k]
                                        if k == 0:
                                            prompt_tokens.append(solution["usage"]["prompt_tokens"])
                                        completion_tokens.append(solution["usage"]["completion_tokens"])
                                    for ii, logs in enumerate(logs_list):
                                        if ii == 0:
                                            prompt_tokens.append(logs[j]["record"]["choose"]["usage"]["prompt_tokens"])
                                        completion_tokens.append(logs[j]["record"]["choose"]["usage"]["completion_tokens"])
                                        
                            elif args.reasoning == "SBP":
                                output_keys = [parse_answer(args, log[j]['record']['solution']["output"]) for log in logs_list]
                                output_keys = [output for output in output_keys if output != None]
                                if len(output_keys) == 0:
                                    output_keys = [None]
                                if count == 0:
                                    prompt_tokens = []
                                    completion_tokens = []
                                    for k in range(0, len(logs_list)):
                                        log = logs_list[k][j]
                                        if k == 0:
                                            prompt_tokens.append(log["record"]["principles"]["usage"]["prompt_tokens"])
                                        completion_tokens.append(log["record"]["principles"]["usage"]["completion_tokens"])
                                        prompt_tokens.append(log["record"]["solution"]["usage"]["prompt_tokens"])
                                        completion_tokens.append(log["record"]["solution"]["usage"]["completion_tokens"])
                                        
                            elif args.reasoning == "MAD":
                                output_keys = [parse_answer(args, log["output"]) for log in logs[j]["record"][f"round{m+1}"]]
                                output_keys = [output for output in output_keys if output != None]
                                if len(output_keys) == 0:
                                    output_keys = [None]
                                if count == 0:
                                    prompt_tokens = []
                                    completion_tokens = []
                                    for k in range(0, m+1):
                                        prompt_tokens += [log["usage"]["prompt_tokens"] for log in logs[j]["record"][f"round{k+1}"]]
                                        completion_tokens += [log["usage"]["completion_tokens"] for log in logs[j]["record"][f"round{k+1}"]]
                            if count == 0:   
                                prompt_tokens_ += sum(prompt_tokens)     
                                completion_tokens_ += sum(completion_tokens)
                            
                            most_common_elements, max_count = find_most_common_elements(output_keys)
                            output_key = random.choice(most_common_elements)
                            
                            if args.dataset in ["GSM8K"]:
                                if output_key != None and abs(float(re.sub(r"[^0-9.-]", "", str(key))) - output_key) < 10**(-4):
                                    acc_num += 1
                            elif args.dataset in ["GSM-Hard"]:
                                if output_key != None and abs(float(key) - output_key) < 10**(-4):
                                    acc_num += 1
                            elif args.dataset in ["MATH", "AIME_2024"]:
                                if is_equiv(str(key), output_key):
                                    acc_num += 1
                            elif args.dataset in ["GPQA"] or "MMLU" in args.dataset:
                                if key == output_key:
                                    acc_num += 1
                        acc = acc_num / l * 100
                        accs.append(acc)
                        
                    acc = sum(accs)/len(accs)
                    total_tokens = prompt_tokens_ + completion_tokens_
                    if args.reasoning in ["ToT"]:
                        cost = args.shot
                    else:   
                        cost = m + 1
                    
    
                    s = f"{args.reasoning.ljust(15)}   " + "{}-shot".format(args.shot).ljust(7) + f"  {str(args.nums).ljust(12)}   prompt_tokens: {str(prompt_tokens_).ljust(10)}  completion_tokens: {str(completion_tokens_).ljust(10)}   cost:" + str("%.12f"%(cost)).ljust(20) + f"tokens: {str(total_tokens).ljust(11)}"  "   Acc: " + "%.4f"%acc
                    print(s)
                    
                    ws_new.append([args.reasoning, subject, str(args.shot), str(args.nums[-1]+1), str(prompt_tokens_), str(completion_tokens_), str("%.15f"%(cost)), str(total_tokens), str("%.4f"%acc)])

            except Exception as e:
                print(args.reasoning, "error")
            
            wb.save(path)

    labels_num = {
        "DiP": 0,
        "CoT": 0,
        "L2M": 0,
        "ToT": 0,
        "S-RF": 0,
        "AnP": 0,
        "SBP": 0,
        "MAD": 0,
    }

    model_name = args.model_name
    sheet = wb.get_sheet_by_name(model_name)

    tokens = []
    accuracy = []
    cost = []
    labels = []
    flag = 0
    for row in sheet.iter_rows(values_only=True):
        if flag == 0:
            flag = 1
            continue
        name = row[0]
        if labels_num[row[0]] >= N_max:
            continue
        labels.append(name)
        labels_num[row[0]] += 1
        tokens.append(int(row[-2]))
        accuracy.append(float(row[-1]))
        cost.append(float(row[-3]))
        
    unique_labels = ["DiP", "CoT", "L2M", "ToT", "S-RF", "SBP", "AnP", "MAD"]     
    wb.close()
    
    plt.figure(figsize=(6, 3))
    ax = plt.gca()
    
    colors = np.array([[5.00000000e-01, 0.00000000e+00, 1.00000000e+00, 1.00000000e+00],
                    [2.17647059e-01, 4.29120609e-01, 9.75511968e-01, 1.00000000e+00],
                    [17.25490196e-02, 8.82927610e-01, 1, 1.00000000e+00],
                    [3.54901961e-01, 9.74138602e-01, 7.82927610e-01, 1.00000000e+00],
                    [6.45098039e-01, 9.74138602e-01, 6.22112817e-01, 1.00000000e+00],
                    [1, 7.82927610e-01, 5.34676422e-01, 1.00000000e+00],
                    [1.00000000e+00, 4.29120609e-01, 2.19946358e-01, 1.00000000e+00],
                    [1.00000000e+00, 1.22464680e-16, 6.12323400e-17, 1.00000000e+00],])
    
    for i, label in enumerate(unique_labels):
        mask = [label == lb for lb in labels]  
        plt.scatter([cost[idx] for idx, val in enumerate(mask) if val],  
                    [accuracy[idx] for idx, val in enumerate(mask) if val],  
                    color=colors[i], label=label, s=36, marker=marker_dict[label], edgecolor='black', linewidths=0.75, zorder=5)  
        plt.plot([cost[idx] for idx, val in enumerate(mask) if val],  
                [accuracy[idx] for idx, val in enumerate(mask) if val], 
                color=colors[i], linewidth=1.5, marker=marker_dict[label])
    
    x_ticks = sampling_times
    y_max = max(accuracy) 
    y_min = min(accuracy) 
    ind_line = (y_max - y_min) / 7.5
    plt.ylim(bottom=None, top=y_max + ind_line * 1.5)
    
    def custom_formatter(y, pos):
        if y > max(accuracy) + 0.9:  
            return ""
        else:
            return f"{int(y)}" 
            
    for x in x_ticks:
        plt.axvline(x=x, color='gray', linestyle='--', alpha=0.5, zorder=0)  
        mask = [cost[idx] == x for idx in range(len(cost))]
        if any(mask):
            best_idx = np.argmax([accuracy[idx] for idx, val in enumerate(mask) if val])
            best_label = [labels[idx] for idx, val in enumerate(mask) if val][best_idx]
            plt.scatter(x, y_max + ind_line, color=colors[unique_labels.index(best_label)], 
                        marker=marker_dict[best_label], s=64, zorder=8, edgecolor='black', linewidths=1)

    plt.axhline(y=y_max + ind_line / 2, color='k', linestyle='-', alpha=0.5, zorder=10, linewidth=1.5)
    plt.rcParams['xtick.labelsize'] = 14
    plt.rcParams['ytick.labelsize'] = 14
    plt.rcParams['font.weight'] = 'bold'
    plt.rcParams['axes.labelweight'] = 'bold'
    plt.rcParams['axes.titleweight'] = 'bold'
    plt.title(f"{model_names_formal[args.model_name]}", fontsize=18)
    plt.xlabel('Sampling Time', fontsize=14, fontweight='bold')
    plt.ylabel('Accuracy', fontsize=14, fontweight='bold')
    
    plt.xticks(x_ticks)
    yticks = ax.get_yticks()
    yticks = yticks[yticks <= max(accuracy)]

    ax.set_yticks(yticks)
    ax.set_yticklabels([str(int(tick)) for tick in yticks])
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))
    ax.yaxis.set_major_formatter(FuncFormatter(custom_formatter))
    
    plt.text(pos, y_max + ind_line,  r'$\mathbf{P}_{N}^*$', fontsize=12, ha='center', va='center')  
    plt.rcParams['font.weight'] = 'bold'
    plt.rcParams['axes.labelweight'] = 'bold'
    plt.rcParams['axes.titleweight'] = 'bold'
    
    plt.legend(loc='best', fontsize=12, framealpha=0.9, ncol=2, markerscale=1.5, handletextpad=0.5, columnspacing=1.0)

    plt.savefig(os.path.join(log_path, args.dataset, args.model_name, "pics", f"Performance_N.png"), bbox_inches='tight', dpi = 600)
    plt.close()
    print(f"Performance_N.png saved to {os.path.join(log_path, args.dataset, args.model_name, 'pics', 'Performance_N.png')}")
